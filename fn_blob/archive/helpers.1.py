from . import metadata
import openpyxl as xlsx
import xlrd as xls
from typing import List
import io

class WorkbookHelper:
    _wb: Union[xlsx.Workbook, xls.Book]
    _wb_type: str
    _sheets: List[metadata.SheetMetadata] = []
    def __init__(self, xl_bytes = io.BufferedIOBase):
        try:
            _wb = xlsx.load_workbook(filename=xl_bytes)
            _wb_type = "xlsx"
            #import methods
        except:
            try:
            _wb = xls.open_workbook(file_contents=xl_bytes)
            _wb_type = "xls"
            #import methods
            except:
                _wb = None
                _wb_type = "err"
    
    def load_sheets(self):
        _sheets = [metadata.SheetMetadata(s) for s in self._wb.sheetnames]
        return self._sheets


class TemplateHelper(WorkbookHelper):
    templates: List[metadata.TemplateMetadata] = []
    def __init__(self, xl_bytes):
        super().__init__(xl_bytes)

    def load_templates(self):
        for s in self._wb.sheetnames:
            ws = self.load_sheet(s)
            t = metadata.TemplateMetadata()
            t.name = s
            t.top_left = (ws.min_column, ws.min_row)
            t.bottom_right = (ws.max_column, ws.max_row)
            self.templates = self.templates + [t]
    
    def load_sheet(self, sheetname: str):
        return self._wb.get_sheet_by_name(sheetname)

