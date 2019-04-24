from . import utility
import openpyxl as xlsx
import xlrd as xls
from typing import List, Dict, Iterable, Union, Any
import io
import pyexcel as pyxl

class WorkBookRepository:
    def load_workbook(self, xl_bytes: io.BufferedIOBase, extension: str):
        pass
    def get_sheet_names(self) -> List[str]:
        pass
    def get_sheet_data(self, sheet_name: str, range: str) -> Iterable:
        pass

class XlsxRepository(WorkBookRepository):
    _wb: xlsx.Workbook
    def load_workbook(self, xl_bytes, extension):
        _wb = xlsx.load_workbook(filename=xl_bytes)

    def get_sheet_names(self):
        return self._wb.sheetnames

    def get_sheet_data(self, sheet_name: str, range: str):
        ws = self._wb.get_sheet_by_name(sheet_name)
        cells = []
        if ":" in range:
            range_arr = range.split(":") 
            cells = ws[range_arr[0]:range_arr[1]]
        else:
            cells = ws[range]

        for cell in cells:
            yield {
                "row": cell.row -1,
                "col": cell.column -1,
                "cell": f"{utility.number_to_char(cell.row -1)}{cell.column -1}",
                "value": cell.value
            }

class XlsRepository(WorkBookRepository):
    _wb: xls.Book
    def load_workbook(self, xl_bytes, extension):
        self._wb = xls.open_workbook(file_contents=xl_bytes.read())

    def get_sheet_names(self):
        return self._wb.sheet_names()

    def get_sheet_data(self, sheet_name: str, range: str) -> Iterable:
        for coords in utility.cell_to_coords(range):
            yield {
                "row": coords[0],
                "col": coords[1],
                "cell": f"{utility.number_to_char(coords[1])}{coords[0]}",
                "value": self._wb.sheet_by_name(sheet_name).cell_value(*coords)
            }

class PyxlRepository(WorkBookRepository):
    _wb: pyxl.Book
    def load_workbook(self, xl_bytes, extension: str):
        self._wb = pyxl.get_book(file_stream=xl_bytes, file_type=extension)

    def get_sheet_names(self):
        return self._wb.to_dict().keys()

    def get_sheet_data(self, sheet_name: str, range: str) -> Iterable:
        for coords in utility.cell_to_coords(range):
            yield {
                "row": coords[0],
                "col": coords[1],
                "cell": f"{utility.number_to_char(coords[1])}{coords[0]}",
                "value": self._wb[sheet_name][coords[0],coords[1]]
            }

repositories = [PyxlRepository(), XlsRepository(), XlsxRepository()]